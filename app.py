import streamlit as st
import pandas as pd
import sqlite3
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

st.set_page_config(page_title="Sistema de Cubagem", layout="wide")

DB_NAME = "faturamento.db"

def init_db():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    # Apaga o banco antigo para recriar com a nova coluna "rota_pai" que controla as mesclagens
    c.execute('DROP TABLE IF EXISTS espelho')
    c.execute('''
        CREATE TABLE espelho (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cidades_rotas TEXT,
            cubagem TEXT,
            lote TEXT,
            romaneio TEXT,
            box TEXT,
            horario_nf TEXT,
            transp TEXT,
            is_header BOOLEAN,
            rota_pai TEXT
        )
    ''')
    conn.commit()
    conn.close()

def carregar_dados_do_db():
    # Adicionada proteção para caso o banco não exista ainda
    try:
        conn = sqlite3.connect(DB_NAME)
        df = pd.read_sql_query("SELECT * FROM espelho", conn)
        conn.close()
        return df
    except:
        return pd.DataFrame()

def limpar_banco():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("DELETE FROM espelho")
    conn.commit()
    conn.close()

def formatar_rota(nome_rota):
    if isinstance(nome_rota, str):
        return nome_rota.replace('AZ ', 'AZUL ').replace('VM ', 'VERMELHA ')
    return nome_rota

def extrair_ax(cidade_str):
    if pd.isna(cidade_str) or not isinstance(cidade_str, str):
        return None, None
    cidade_sem_cubagem = str(cidade_str).split('/')[0]
    ax = cidade_sem_cubagem.split('-')[0].strip()
    return ax, cidade_sem_cubagem

def processar_arquivos(cubagem_file, lotes_file):
    df_cubagem = pd.read_csv(cubagem_file) if cubagem_file.name.endswith('.csv') else pd.read_excel(cubagem_file)
    df_lotes = pd.read_csv(lotes_file) if lotes_file.name.endswith('.csv') else pd.read_excel(lotes_file)

    df_lotes.columns = df_lotes.columns.astype(str).str.lower().str.strip()
    dict_lotes = {}
    if 'ax' in df_lotes.columns and 'numlote' in df_lotes.columns:
        df_lotes['ax'] = df_lotes['ax'].astype(str).str.strip()
        dict_lotes = dict(zip(df_lotes['ax'], df_lotes['numlote']))

    if 'rotas' in df_cubagem.columns:
        df_cubagem = df_cubagem.dropna(subset=['rotas'])

    dados_processados = []

    for index, row in df_cubagem.iterrows():
        rota = formatar_rota(row.get('rotas', ''))
        transportadora = row.get('transportadora', '')
        
        primeira_cidade = ""
        if 'filial1/cubagem' in row and pd.notna(row['filial1/cubagem']):
            _, primeira_cidade = extrair_ax(row['filial1/cubagem'])

        # Adiciona a linha da rota em branco (is_header = True) e grava o "rota_pai"
        dados_processados.append((rota, "", "", "", primeira_cidade, "", "", True, rota))

        for i in range(1, 13):
            col_nome = f'filial{i}/cubagem'
            if col_nome in row and pd.notna(row[col_nome]) and str(row[col_nome]).strip() != '':
                ax, nome_cidade = extrair_ax(row[col_nome])
                lote_encontrado = dict_lotes.get(str(ax), "")
                # Adiciona as cidades (is_header = False) com vínculo na mesma "rota_pai"
                dados_processados.append((nome_cidade, "", lote_encontrado, "", "", "", transportadora, False, rota))

    limpar_banco()
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.executemany('''
        INSERT INTO espelho (cidades_rotas, cubagem, lote, romaneio, box, horario_nf, transp, is_header, rota_pai)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', dados_processados)
    conn.commit()
    conn.close()

def gerar_excel(df_atual):
    wb = Workbook()
    ws = wb.active
    ws.title = "Espelho Carregamento"

    ws.sheet_view.showGridLines = False # Remove a grade do fundo

    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    no_border = Border()

    # NOVOS CABEÇALHOS
    headers = ["ROTAS 1 BATIDA", "CONFERENTE", "LOTE", "ROMANEIO", "1° FILIAL", "N° NOTA FISCAL", "TRANSPORTADORA"]
    ws.append(headers)
    
    for col in range(1, 8):
        cell = ws.cell(row=1, column=col)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col)].width = 25

    for row_idx, row in df_atual.iterrows():
        excel_row = row_idx + 2
        is_header = row['is_header']
        
        valores = [row['cidades_rotas'], row['cubagem'], row['lote'], row['romaneio'], row['box'], row['horario_nf'], row['transp']]
        
        for col_idx, val in enumerate(valores):
            cell = ws.cell(row=excel_row, column=col_idx+1, value=val)
            cell.alignment = center_align
            
            if is_header:
                cell.font = bold_font
                cell.fill = white_fill
                cell.border = no_border 
            else:
                cell.border = thin_border

    # --- LÓGICA DE MESCLAR AS CÉLULAS DAS CIDADES NO EXCEL ---
    route_start_row = None
    for row_idx, row in df_atual.iterrows():
        excel_row = row_idx + 2
        is_header = row['is_header']
        
        if is_header:
            # Se já tinha uma rota acontecendo antes, mescla ela antes de iniciar a nova
            if route_start_row is not None and (excel_row - 1) > route_start_row:
                for col in [2, 4, 6, 7]: # Colunas: Conferente, Romaneio, NF, Transportadora
                    ws.merge_cells(start_row=route_start_row, start_column=col, end_row=excel_row - 1, end_column=col)
            # Define o início das cidades da nova rota
            route_start_row = excel_row + 1 
            
    # Mescla a última rota da planilha
    ultimo_row = len(df_atual) + 1
    if route_start_row is not None and ultimo_row > route_start_row:
        for col in [2, 4, 6, 7]:
            ws.merge_cells(start_row=route_start_row, start_column=col, end_row=ultimo_row, end_column=col)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# Inicializa o banco ao rodar
init_db()

st.title("🚛 Plataforma de Faturamento e Cubagem")
st.markdown("Preencha os dados do carregamento. As alterações feitas na tabela são salvas automaticamente.")

with st.sidebar:
    st.header("1. Carregar Dados")
    st.info("Faça o upload dos arquivos para iniciar um novo espelho.")
    cubagem_file = st.file_uploader("Planilha de Cubagem", type=['xlsx', 'xls', 'csv'])
    lotes_file = st.file_uploader("Planilha Detalhes (82)", type=['xlsx', 'xls', 'csv'])
    
    if st.button("Processar Novas Planilhas", use_container_width=True, type="primary"):
        if cubagem_file and lotes_file:
            processar_arquivos(cubagem_file, lotes_file)
            st.success("Dados processados com sucesso!")
            st.rerun()
        else:
            st.warning("Por favor, anexe as duas planilhas.")

df_tela = carregar_dados_do_db()

if not df_tela.empty:
    col1, col2 = st.columns([8, 2])
    with col1:
        st.write("Dê dois cliques nas colunas editáveis de uma cidade. **A digitação vai preencher todas as cidades da rota automaticamente!**")
    with col2:
        if st.button("🔄 Sincronizar (Ver edições)"):
            st.rerun()

    def colorir_linha_rota(row):
        if row['is_header']:
            return ['background-color: #FFFFFF; color: #000000; font-weight: bold; border: none !important;'] * len(row)
        return [''] * len(row)
    
    df_estilizado = df_tela.style.apply(colorir_linha_rota, axis=1)

    # NOVOS CABEÇALHOS NA INTERFACE
    column_config = {
        "id": None, 
        "is_header": None, 
        "rota_pai": None,
        "cidades_rotas": st.column_config.TextColumn("ROTAS / CIDADES", disabled=True),
        "cubagem": st.column_config.TextColumn("CONFERENTE"),
        "lote": st.column_config.TextColumn("LOTE", disabled=True),
        "romaneio": st.column_config.TextColumn("ROMANEIO"),
        "box": st.column_config.TextColumn("1° FILIAL", disabled=True),
        "horario_nf": st.column_config.TextColumn("N° NOTA FISCAL"),
        "transp": st.column_config.TextColumn("TRANSPORTADORA", disabled=True)
    }

    tabela_editada = st.data_editor(
        df_estilizado, 
        column_config=column_config,
        use_container_width=True,
        hide_index=True,
        key="editor_faturamento"
    )

    mudancas = st.session_state["editor_faturamento"]
    if mudancas.get("edited_rows"):
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        
        teve_erro_edicao = False
        
        for row_index, alteracoes in mudancas["edited_rows"].items():
            is_header = df_tela.iloc[row_index]['is_header']
            
            if is_header:
                teve_erro_edicao = True
            else:
                # LÓGICA DE MESCLAGEM NA WEB: Atualiza TODAS as cidades que pertencem à mesma Rota
                rota_pai = df_tela.iloc[row_index]['rota_pai']
                for coluna, novo_valor in alteracoes.items():
                    c.execute(f"UPDATE espelho SET {coluna} = ? WHERE rota_pai = ? AND is_header = 0", (str(novo_valor), rota_pai))
                    
        conn.commit()
        conn.close()
        
        # Atualiza a página para mostrar a "mágica" preenchendo as outras linhas
        st.rerun() 

    st.divider()

    excel_data = gerar_excel(df_tela)
    st.download_button(
        label="📥 2. Baixar Espelho Final (Excel)",
        data=excel_data,
        file_name="Relatorio_Carregamento_Final.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary"
    )
else:
    st.info("Nenhum dado carregado. Use a barra lateral para fazer o upload das planilhas de Cubagem e Lotes.")
